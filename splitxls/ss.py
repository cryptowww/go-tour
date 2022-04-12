from openpyxl import load_workbook,Workbook


wb = load_workbook("2021年7月内部考核版直接费用追踪.xlsx")
sheetnames = wb.sheetnames

for name in sheetnames:
    ws = wb.get_sheet_by_name(name)
    print(ws)
    # 创建新的Excel
    wb2 = Workbook()
    # 获取当前sheet
    ws2 = wb2.active
    #两个for循环遍历整个excel的单元格内容
    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            # 写入新Excel
            ws2.cell(row=i+1, column=j+1, value=cell.value)
            # 设置新Sheet的名称
            ws2.title = name

    wb2.save(name + ".xlsx")

